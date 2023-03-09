from . import forms
from .forms import ProductForm, ClientForm, LivraisonForm, TransporteurForm , Niveau1Form, Niveau2Form, Niveau3Form, Niveau4Form, Niveau5Form
from django.shortcuts import redirect, render, get_object_or_404
from django.contrib.auth.decorators import login_required
from  django.views.decorators.cache import cache_control 
from  django.contrib import messages
from django.http import HttpResponseRedirect, HttpResponse, FileResponse, JsonResponse
from django.core.paginator import Paginator
from django.db.models import Count, Sum
from stock.models import *
from escpos.printer import Usb
from django.db import transaction, models
from django.utils import timezone
import uuid
import csv
import openpyxl
import datetime
from datetime import date
from django.views.generic import View





@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def index(request):
    articles = Product.objects.all().count()
    articles_commandes = Commande.objects.all().count()
    articles_livres = Livraison.objects.filter(livre=True).count()
    context={"articles":articles, "articles_commandes":articles_commandes, "articles_livres":articles_livres}
    return render(request, 'stock/index.html', context)



@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def client(request):
    clients = Client.objects.all().order_by('-id')
    paginator = Paginator(clients, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context={"page_obj":page_obj}
    return render(request, 'stock/client.html', context)  


@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def add_client(request):
    if request.method == 'POST':
        form = ClientForm(request.POST)
        if form.is_valid():
            form.save()
            cont_type1 = form.cleaned_data.get('rasion_sociale')
            cont_type2 = form.cleaned_data.get('rasion_sociale')
            messages.success(request, f"{cont_type1}, {cont_type2} ajouté avec succès!")
            return HttpResponseRedirect('client')
        else:
            messages.error(request, "Veuillez verifiez svpla saisie!")
            return render(request, 'stock/add_client.html', {'form':form})
    else:
        form = ClientForm()
    return render(request, 'stock/add_client.html', {'form':form})


@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def edit_client(request, id):
    client = Client.objects.get(id=id)
    if request.method == 'POST':
        form = ClientForm(request.POST,request.FILES, instance=client)
        if form.is_valid():
            form.save(id)
            messages.success(request, f"Mise à jour {client.rasion_sociale} ok !")
            return redirect('client')
    else:
        form = ClientForm(instance=client)
    return render(request, 'stock/edit_client.html', {'form':form, "client":client})



@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def delete_client(request, id):
    client = Client.objects.get(id=id)
    if request.method=='POST':
        client.delete()
        messages.success(request, f'{client.rasion_sociale} supprimé !')
        return redirect("client")
    return render(request, 'stock/delete_client.html', {"client":client})

@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def transport(request):
    transports = Transporteur.objects.all().order_by('-id')
    paginator = Paginator(transports, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context={"page_obj":page_obj}
    return render(request, 'stock/transport.html', context)  



@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def add_transport(request):
    if request.method == 'POST':
        form = TransporteurForm(request.POST)
        if form.is_valid():
            form.save()
            cont_type1 = form.cleaned_data.get('name_transp')
            messages.success(request, f"{cont_type1} ajouté avec succès!")
            return HttpResponseRedirect('transport')
        else:
            messages.error(request, "Veuillez verifiez svpla saisie!")
            return render(request, 'stock/add_client.html', {'form':form})
    else:
        form = TransporteurForm()
    return render(request, 'stock/add_transport.html', {'form':form})




@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)  
def edit_transport(request, id):
    transport = Transporteur.objects.get(id=id)
    if request.method == 'POST':
        form = TransporteurForm(request.POST,request.FILES, instance=transport)
        if form.is_valid():
            form.save(id)
            messages.success(request, f"Mise à jour {transport.name_transp} ok !")
            return redirect('transport')
    else:
        form = TransporteurForm(instance=transport)
    return render(request, 'stock/edit_transport.html', {'form':form, "transport":transport})



@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def delete_transport(request, id):
    transport = Transporteur.objects.get(id=id)
    if request.method=='POST':
        transport.delete()
        messages.success(request, f'{transport.name_transp} supprimé !')
        return redirect("transport")
    return render(request, 'stock/delete_transport.html', {"transport":transport})




@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def niveau(request):
    return render(request, 'stock/niveau.html')

@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def create_n1(request):
    niveau1 = Niveau1.objects.all()
    if request.method == 'POST':
        form = Niveau1Form(request.POST)
        if form.is_valid():
            form.save()
            cont_type = form.cleaned_data.get('name')
            messages.success(request, f"{cont_type} ajouté avec succès!")
            return HttpResponseRedirect('create_n1')
        else:
            messages.error(request, "Veuillez verifiez svpla saisie!")
            return render(request, 'stock/create_n1.html', {"niveau1":niveau1, 'form':form})
    else:
        form = Niveau1Form()
    return render(request, 'stock/create_n1.html', {"niveau1":niveau1, 'form':form})



@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)  
def edit_n1(request, id):
    niveau = Niveau1.objects.get(id=id)
    if request.method == 'POST':
        form = Niveau1Form(request.POST,request.FILES, instance=niveau)
        if form.is_valid():
            form.save(id)
            cont_type = form.cleaned_data.get('name')
            messages.success(request, f"{cont_type} mis à jour avec succès!")
            return redirect('create_n1')
    else:
        form = Niveau1Form(instance=niveau)
    return render(request, 'stock/edit_n1.html', {"niveau":niveau, 'form':form})


@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def create_n2(request):
    niveau2 = Niveau2.objects.all()
    if request.method == 'POST':
        form = Niveau2Form(request.POST)
        if form.is_valid():
            form.save()
            cont_type = form.cleaned_data.get('name')
            messages.success(request, f"{cont_type} ajouté avec succès!")
            return HttpResponseRedirect('create_n2')
        else:
            messages.error(request, "Veuillez verifiez svpla saisie!")
            return render(request, 'stock/create_n2.html', {"niveau2":niveau2, 'form':form})
    else:
        form = Niveau2Form()
    return render(request, 'stock/create_n2.html', {"niveau2":niveau2, 'form':form})



@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)  
def edit_n2(request, id):
    niveau2 = Niveau2.objects.get(id=id)
    if request.method == 'POST':
        form = Niveau2Form(request.POST,request.FILES, instance=niveau2)
        if form.is_valid():
            form.save(id)
            cont_type = form.cleaned_data.get('name')
            messages.success(request, f"{cont_type} mis à jour avec succès!")
            return redirect('create_n2')
    else:
        form = Niveau2Form(instance=niveau2)
    return render(request, 'stock/edit_n2.html', {"niveau2":niveau2, 'form':form})



@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def create_n3(request):
    niveau3 = Niveau3.objects.all()
    if request.method == 'POST':
        form = Niveau3Form(request.POST)
        if form.is_valid():
            form.save()
            cont_type = form.cleaned_data.get('name')
            messages.success(request, f"{cont_type} ajouté avec succès!")
            return HttpResponseRedirect('create_n3')
        else:
            messages.error(request, "Veuillez verifiez svpla saisie!")
            return render(request, 'stock/create_n3.html', {"niveau3":niveau3, 'form':form})
    else:
        form = Niveau3Form()
    return render(request, 'stock/create_n3.html', {"niveau3":niveau3, 'form':form})


@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)  
def edit_n3(request, id):
    niveau3 = Niveau3.objects.get(id=id)
    if request.method == 'POST':
        form = Niveau3Form(request.POST,request.FILES, instance=niveau3)
        if form.is_valid():
            form.save(id)
            cont_type = form.cleaned_data.get('name')
            messages.success(request, f"{cont_type} mis à jour avec succès!")
            return redirect('create_n3')
    else:
        form = Niveau3Form(instance=niveau3)
    return render(request, 'stock/edit_n3.html', {"niveau3":niveau3, 'form':form})




@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def create_n4(request):
    niveau4 = Niveau4.objects.all()
    if request.method == 'POST':
        form = Niveau4Form(request.POST)
        if form.is_valid():
            form.save()
            cont_type = form.cleaned_data.get('name')
            messages.success(request, f"{cont_type} ajouté avec succès!")
            return HttpResponseRedirect('create_n4')
        else:
            messages.error(request, "Veuillez verifiez svpla saisie!")
            return render(request, 'stock/create_n4.html', {"niveau4":niveau4, 'form':form})
    else:
        form = Niveau4Form()
    return render(request, 'stock/create_n4.html', {"niveau4":niveau4, 'form':form})


@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)  
def edit_n4(request, id):
    niveau4 = Niveau4.objects.get(id=id)
    if request.method == 'POST':
        form = Niveau4Form(request.POST,request.FILES, instance=niveau4)
        if form.is_valid():
            form.save(id)
            cont_type = form.cleaned_data.get('name')
            messages.success(request, f"{cont_type} mis à jour avec succès!")
            return redirect('create_n4')
    else:
        form = Niveau4Form(instance=niveau4)
    return render(request, 'stock/edit_n4.html', {"niveau4":niveau4, 'form':form})




@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def create_n5(request):
    niveau5 = Niveau5.objects.all()
    if request.method == 'POST':
        form = Niveau5Form(request.POST)
        if form.is_valid():
            form.save()
            cont_type = form.cleaned_data.get('name')
            messages.success(request, f"{cont_type} ajouté avec succès!")
            return HttpResponseRedirect('create_n5')
        else:
            messages.error(request, "Veuillez verifiez svpla saisie!")
            return render(request, 'stock/create_n5.html', {"niveau5":niveau5, 'form':form})
    else:
        form = Niveau5Form()
    return render(request, 'stock/create_n5.html', {"niveau5":niveau5, 'form':form})


@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)  
def edit_n5(request, id):
    niveau5 = Niveau5.objects.get(id=id)
    if request.method == 'POST':
        form = Niveau5Form(request.POST,request.FILES, instance=niveau5)
        if form.is_valid():
            form.save(id)
            cont_type = form.cleaned_data.get('name')
            messages.success(request, f"{cont_type} mis à jour avec succès!")
            return redirect('create_n5')
    else:
        form = Niveau5Form(instance=niveau5)
    return render(request, 'stock/edit_n5.html', {"niveau5":niveau5, 'form':form})




@login_required 
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def commande(request):
    if request.method == 'POST':
        supplier_id = request.POST.get('fournisseur')
        supplier = Client.objects.get(pk=supplier_id)
        adresse_livraison = request.POST.get('adresse_livraison')
        date_commande = request.POST.get('date_commande')
        num_commande = request.POST.get('num_commande')
        transport_id = request.POST.get('transport')
        transport = Transporteur.objects.get(pk=transport_id)
        commande = Commande.objects.create(client=supplier,
                                            date_commande=date_commande,
                                            adresse_livraison=adresse_livraison,
                                            num_commande=num_commande,
                                            transport=transport
                                            )
        selected_products = request.POST.getlist('articles')
        number = 0
        with transaction.atomic():
            for article_id in selected_products:
                number += 1
                quantity = int(request.POST.get(f'quantity-{number}'))
                article = Product.objects.get(id=article_id)
                # Vérifiez si la quantité demandée est disponible
                if article.stock >= quantity:
                    article.stock = models.F('stock') - quantity
                    article.save()

                    LigneCommande.objects.create(commande=commande,
                                                     article=article,
                                                     quantite=quantity)

                else:
                        # Si la quantité n'est pas disponible, annulez la commande et affichez un message d'erreur
                    commande.delete()
                    messages.error(request, f"La quantité demandée pour '{article.name}' n'est pas disponible.")

                    return redirect('livraison')

            Livraison.objects.create(commande=commande,
                                      num_livraison=commande.num_commande,
                                      livre=False,
                                      )

        messages.success(request, 'Votre commande a été passée avec succès!')
        return redirect('livraison')

    else:
        articles = Product.objects.all()
        suppliers = Client.objects.all()
        transports = Transporteur.objects.all()
        prefixe = 'CMD-'
        idc = f'{Commande.objects.count()+1:05d}'
        suffixe = date.today()
        sep = '-'
        num_commande = prefixe + idc + sep  + str(suffixe)
        context = {'articles': articles, "suppliers":suppliers, "transports":transports, "num_commande":num_commande}
        return render(request, 'stock/commande.html', context)




@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def print_commande(request, pk):
    commande_print = Livraison.objects.get(pk=pk)
    com_id = commande_print.id
    ma_commande = get_object_or_404(Commande, id=com_id)
    article_commandes = LigneCommande.objects.filter(commande=ma_commande)
    nbr_items = article_commandes.count()
    context={"article_commandes":article_commandes, "ma_commande":ma_commande, "nbr_items":nbr_items}
    return render(request, 'stock/print_commande.html', context)



@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def print_commande2(request, pk):
    commande_print = Livraison.objects.get(pk=pk)
    com_id = commande_print.id
    ma_commande = get_object_or_404(Commande, id=com_id)
    article_commandes = LigneCommande.objects.filter(commande=ma_commande)
    context={"article_commandes":article_commandes, "ma_commande":ma_commande}
    return render(request, 'stock/print_commande2.html', context)



@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def livraison(request):
    livraisons = Livraison.objects.filter(livre=False).order_by('-id')
    paginator = Paginator(livraisons, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context={"page_obj":page_obj}
    return render(request, 'stock/livraison.html', context)


@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def edit_livraison(request, livraison_id):
    livraison = get_object_or_404(Livraison, id=livraison_id)
    commande = livraison.commande
    form = LivraisonForm(request.POST or None, instance=livraison)
    if form.is_valid():
        livraison = form.save(commit=False)
        commande.save()
        livraison.livre = True
        livraison.save()
        messages.success(request, 'Votre commande a été livré avec succès!')
        return redirect('suivi')
    return render(request, 'stock/edit_livraison.html', {'form': form, 'livraison': livraison})



@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def edit_livraison2(request, livraison_id):
    livraison = get_object_or_404(Livraison, id=livraison_id)
    return render(request, 'stock/edit_livraison2.html', {'livraison': livraison})





@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def produit(request):
    articles = Product.objects.all().order_by('-id')
    paginator = Paginator(articles, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context={"page_obj":page_obj}
    return render(request, 'stock/produit.html', context)



def stock_barre(request):
    barcode = request.GET.get('barcode')
    if barcode:
        product = get_object_or_404(Product, barcode=barcode)
        data = {
            'name': product.name,
            'code': product.code,
            'stock': product.stock,
            'description': product.description,
            'sous_contenaire': str(product.sous_contenaire),
            'barcode': product.barcode.url if product.barcode else '',
        }
        return JsonResponse(data)
    return render(request, 'stock/stock_barre.html')



@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def add_product(request):
    year = datetime.datetime.now().year
    code = f'ARTI-{Product.objects.count()+1:09d}-{year}'
    if request.method=="POST":
        form = ProductForm(request.POST)
        if form.is_valid():
            product = form.save(commit=False)
            product.code = code
            product.save()
            pro_code = form.cleaned_data.get('code')
            nom_pro = form.cleaned_data.get('name')
            messages.success(request, f"L'élément {nom_pro} code {pro_code}  ajouté avec succes !")
            return HttpResponseRedirect('produit')
        else:
            messages.error(request, "Veuillez verifiez svp l'article existe déja!!")
            return render(request, 'stock/add_product.html', {"form":form})
    else:
        form = ProductForm(initial={'code': code})
        return render(request, 'stock/add_product.html', {"form":form})
    
@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def edit_product(request, id):
    product = get_object_or_404(Product, id=id)
    if request.method == 'POST':
        form = ProductForm(request.POST,request.FILES, instance=product)
        if form.is_valid():
            form.save(id)
            messages.success(request, f"Mise à jour {product.code} ok !")
            return redirect('produit')
    else:
        form = ProductForm(instance=product)
    return render(request, 'stock/edit_product.html', {'form':form})




@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True) 
def delete_product(request, id):
    product = Product.objects.get(id=id)
    if request.method=='POST':
        product.delete()
        messages.success(request, f'{product.code} supprimé !')
        return redirect("produit")
    return render(request, 'stock/delete_product.html', {"product":product})



class SearchProductView(View):
    def get(self, request, *args, **kwargs):
        code = request.GET.get('code')
        products = Product.objects.filter(code__icontains=code)
        data = {
            'products': list(products.values('name'))
        }
        return JsonResponse(data)



    
@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def stock_in(request):
    operations = Operation.objects.all().order_by('-id')
    paginator = Paginator(operations, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context={"page_obj":page_obj}
    return render(request, 'stock/stock_in.html', context)


@login_required 
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def add_stock(request):
    if request.method == 'POST':
        supplier_id = request.POST.get('fournisseur')
        supplier = Client.objects.get(pk=supplier_id)
        operation = Operation.objects.create(fournisseur=supplier
                                           )
        number = 0
        selected_products = request.POST.getlist('products')
        with transaction.atomic():
            for product_id in selected_products:
                number += 1
                quantity = int(request.POST.get(f'quantity-{number}'))
                product = Product.objects.get(id=product_id)
                # Vérifiez si la quantité demandée est disponible
                if quantity >= 1:
                    product.stock = models.F('stock') + quantity
                    product.save()
                    LigneOperation.objects.create(operation=operation,
                                                     product=product,
                                                     quantity=quantity)
                   
                else:
                        # Si la quantité n'est pas disponible, annulez la commande et affichez un message d'erreur
                    operation.delete()
                    messages.error(request, f"La quantité {product.code} doit être superieure à au moin 1 ")

                    return redirect('stock_in')

            

        messages.success(request, 'Votre stock à été mise à jour avec succès!')
        return redirect('stock_in')

    else:
        products = Product.objects.all()
        suppliers = Client.objects.all()
        context = {'products': products, 'suppliers': suppliers}
        return render(request, 'stock/add_stock.html', context)
    
    
    
# @login_required
# @cache_control(no_cache=True, must_revalidate=True, no_store=True)
# def add_stock(request):
#     if request.method == 'POST':
#         selected_products = request.POST.getlist('products')
#         for product_id in selected_products:
#             quantity = int(request.POST.get('quantity'))
#             supplier_id = request.POST.get('fournisseur')
#             supplier = Client.objects.get(pk=supplier_id)
#             product = Product.objects.get(pk=product_id)
#             if quantity >= 1:
#                 product.stock += quantity
#                 product.save()
#                 operation = Operation(
#                     products=product,
#                     quantity=quantity,
#                     fournisseur=supplier
#                     )
#                 operation.save()
#                 messages.success(request, f"Le stock {product} mis à jour avec succes !")
#             else:
#                 messages.error(request, f" veuillez verifiez la quantité saisie")
            
#         return redirect('stock_in')

#     products = Product.objects.all()
#     suppliers = Client.objects.all()
#     context = {'products': products, 'suppliers': suppliers}
#     return render(request, 'stock/add_stock.html', context)

 
@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def operation_print(request, id):
    op_id = Operation.objects.get(id=id)
    operations_lignes = LigneOperation.objects.filter(operation_id=op_id)
    context={"operations_lignes":operations_lignes, "op_id":op_id}
    return render(request, 'stock/operation_print.html', context)




@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def suivi(request):
    livraisons = Livraison.objects.all().filter(livre=True).order_by('-id')
    paginator = Paginator(livraisons, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context={"page_obj":page_obj}
    return render(request, 'stock/suivi.html', context)




class ExportProductsCSVView(View):
    
    def get(self, request, *args, **kwargs):
        products = Product.objects.all()
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="products.csv"'

        writer = csv.writer(response)
        writer.writerow(['ID', 'Code', 'Name', 'Stock', 'Niveau', 'Description'])

        for product in products:
            writer.writerow([product.id, product.code, product.name, product.stock, product.sous_contenaire, product.description])

        return response




class ExportProductsExcelView(View):
    
    def get(self, request, *args, **kwargs):
        products = Product.objects.all()

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Products"

        # Écrire les en-têtes de colonne
        worksheet.cell(row=1, column=1, value="ID")
        worksheet.cell(row=1, column=2, value="Code")
        worksheet.cell(row=1, column=3, value="Name")
        worksheet.cell(row=1, column=4, value="Stock")
        worksheet.cell(row=1, column=5, value="Niveau")
        worksheet.cell(row=1, column=6, value="Description")
        # Écrire les données des produits
        row_num = 2
        for product in products:
            worksheet.cell(row=row_num, column=1, value=product.id)
            worksheet.cell(row=row_num, column=2, value=product.code)
            worksheet.cell(row=row_num, column=3, value=product.name)
            worksheet.cell(row=row_num, column=4, value=product.stock)
            worksheet.cell(row=row_num, column=5, value=product.sous_contenaire.name)
            worksheet.cell(row=row_num, column=6, value=product.description)
            row_num += 1
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="products.xlsx"'
        workbook.save(response)
        return response



# @login_required
# @cache_control(no_cache=True, must_revalidate=True, no_store=True)
# def print_label(product):
#     #Crée un objet d'imprimante USB
#     printer = Usb(0x0416, 0x5011, 0, profile="POS-5890")
#     #Configure l'imprimante pour une étiquette de 50 mm de largeur
#     printer.set("CENTER")
#     printer.set("SIZE", 1, 1)
#     printer.set("GAP", 2)
#     #Imprime le nom du produit
#     printer.text(product.name + "\n")
#     #Génère le code-barres et l'imprime
#     barcode_data = generate_barcode(product.number)
#     printer.barcode(barcode_data, "CODE128", 64, 2, '', '')
#     #Imprime le numéro de produit et la quantité
#     printer.text("Product number: " + product.number + "\n")
#     printer.text("Quantity: " + str(product.quantity) + "\n")
#     #Coupe le papier pour terminer l'étiquette
#     printer.cut()
#     #Ferme l'objet d'imprimante USB
#     printer.close()
#     return HttpResponse("Label printed successfully.")
    


# @login_required
# @cache_control(no_cache=True, must_revalidate=True, no_store=True)
# def product_label(request, product_id):
#     #Récupère l'objet Produit correspondant à l'ID donné
#     product = Product.objects.get(pk=product_id)
#     #Imprime une étiquette pour le produit
#     print_label(product)
#     #Redirige l'utilisateur vers la page de détails du produit
#     return redirect('product_detail', product_id=product_id)


    
